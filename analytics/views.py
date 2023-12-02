from datetime import datetime, timedelta

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, F, Value
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseBadRequest
from django.shortcuts import render
from openpyxl import Workbook

from accounts.models import TrafficSource, User
from analytics.models import ReportPermission


@login_required
def request_source_analytics(request):
    report_permissions = ReportPermission.objects.filter(user=request.user)
    if not report_permissions:
        return HttpResponseForbidden('No permission!')

    correct_url = settings.HOST_URL + '/analytics/marketing/reports/download/'
    context = {
        'redirect_url': correct_url
    }
    return render(request, 'datetime_form.html', context)


@login_required
def get_source_analytics(request):
    start_date_str = request.GET.get('start_date', None)
    end_date_str = request.GET.get('end_date', None)

    if start_date_str and end_date_str:

        start_datetime = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M').astimezone()
        end_datetime = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M').astimezone()

        if start_datetime < end_datetime - timedelta(days=30):
            return HttpResponseBadRequest('Report time filter threshold must be less than 30 days')

        q = Q()
        report_permissions = ReportPermission.objects.filter(user=request.user)
        if not report_permissions:
            return HttpResponseForbidden('No permission!')

        for permission in report_permissions:
            q = q | permission.q

        # generate Excel workbook from queryset
        if q is None:
            return HttpResponseBadRequest('There is no data in this period')

        workbook = queryset_to_workbook(
            TrafficSource.objects.filter(
                q,
                created__range=[start_datetime, end_datetime]
            )
        )

        # create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reports.xlsx'

        # write workbook to response
        workbook.save(response)

        return response

    return HttpResponseBadRequest('Invalid data')


def queryset_to_workbook(queryset, sheet_name='Sheet1'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = ['date', 'utm_source', 'utm_medium', 'utm_campaign', 'utm_content', 'utm_term', 'users', 'verified', 'depositors']

    # write headers
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    groups = queryset.values('created__date', 'utm_source', 'utm_medium', 'utm_campaign', 'utm_content', 'utm_term') \
        .annotate(
        user_count=Count('user__id', distinct=True),
        verified_count=Count('user__id', distinct=True, filter=Q(level__gte=User.LEVEL2)),
        depositor_count=Count(
            'user__id', distinct=True,
            filter=Q(user__first_fiat_deposit_date__lte=F('user__date_joined') + Value(timedelta(days=1))) |
                   Q(user__first_crypto_deposit_date__lte=F('user__date_joined') + Value(timedelta(days=1)))
        )
    ).values_list(
        'created__date', 'utm_source', 'utm_medium', 'utm_campaign', 'utm_content', 'utm_term',
        'user_count', 'verified_count', 'depositor_count',
    )

    # write data
    for row_num, row in enumerate(groups, 1):
        for col_num, field_name in enumerate(headers, 1):
            cell = sheet.cell(row=row_num+1, column=col_num)
            cell.value = row[col_num-1]

    return workbook
